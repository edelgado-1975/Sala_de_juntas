from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Sum
from django.utils import timezone
from django.http import HttpResponse
from apps.reservas.models import Reserva
from apps.salas.models import Sala
from apps.usuarios.models import Usuario
import calendar
from datetime import datetime, date, timedelta, time
import io
from django.utils.dateparse import parse_date
from django.utils import timezone

# PDF generation imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# Excel generation imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def es_admin(user):
    return user.es_superusuario or user.es_operativo

@login_required
@user_passes_test(es_admin)
def dashboard_informes(request):
    """
    Vista principal del dashboard de informes y estadísticas.
    """
    total_salas = Sala.objects.count()
    total_usuarios = Usuario.objects.count()
    total_reservas = Reserva.objects.count()
    reservas_hoy = Reserva.objects.filter(
        fecha_inicio__date=timezone.now().date()
    ).count()

    # Estadísticas por sala
    reservas_por_sala = Reserva.objects.values('sala__nombre').annotate(
        total=Count('id')
    ).order_by('-total')

    # Estadísticas por estado
    reservas_por_estado = Reserva.objects.values('estado').annotate(
        total=Count('id')
    )

    context = {
        'total_salas': total_salas,
        'total_usuarios': total_usuarios,
        'total_reservas': total_reservas,
        'reservas_hoy': reservas_hoy,
        'reservas_por_sala': reservas_por_sala,
        'reservas_por_estado': reservas_por_estado,
    }
    return render(request, 'informes/dashboard.html', context)


@login_required
@user_passes_test(es_admin)
def informe_usuario_view(request):
    """Vista para generar informes detallados por usuario."""
    usuarios = Usuario.objects.all().order_by('first_name', 'last_name')
    usuario_id = request.GET.get('usuario_id')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    reservas = Reserva.objects.none()
    usuario_seleccionado = None
    
    if usuario_id:
        usuario_seleccionado = get_object_or_404(Usuario, id=usuario_id)
        reservas = Reserva.objects.filter(usuario=usuario_seleccionado).select_related('sala')
        
        if fecha_inicio_str:
            f_inicio = parse_date(fecha_inicio_str)
            if f_inicio:
                dt_inicio = timezone.make_aware(datetime.combine(f_inicio, time.min))
                reservas = reservas.filter(fecha_inicio__gte=dt_inicio)
        
        if fecha_fin_str:
            f_fin = parse_date(fecha_fin_str)
            if f_fin:
                dt_fin = timezone.make_aware(datetime.combine(f_fin, time.max))
                reservas = reservas.filter(fecha_fin__lte=dt_fin)
            
        reservas = reservas.order_by('-fecha_inicio')

    context = {
        'usuarios': usuarios,
        'reservas': reservas,
        'usuario_seleccionado': usuario_seleccionado,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
    }
    return render(request, 'informes/informe_usuario.html', context)


@login_required
@user_passes_test(es_admin)
def exportar_reservas_pdf(request):
    """Genera un reporte PDF de todas las reservas."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Reservas_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title = Paragraph("<b>SISTEMA DE AGENDAMIENTO SALA DE JUNTAS - SENA</b>", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(f"Reporte de Reservas Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Datos
    # Aplicar filtros si existen
    usuario_id = request.GET.get('usuario_id')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    reservas = Reserva.objects.all().select_related('sala', 'usuario')

    if usuario_id:
        reservas = reservas.filter(usuario_id=usuario_id)
    
    if fecha_inicio_str:
        f_inicio = parse_date(fecha_inicio_str)
        if f_inicio:
            dt_inicio = timezone.make_aware(datetime.combine(f_inicio, time.min))
            reservas = reservas.filter(fecha_inicio__gte=dt_inicio)
            
    if fecha_fin_str:
        f_fin = parse_date(fecha_fin_str)
        if f_fin:
            dt_fin = timezone.make_aware(datetime.combine(f_fin, time.max))
            reservas = reservas.filter(fecha_fin__lte=dt_fin)

    reservas = reservas.order_by('-fecha_inicio')
    
    data = [['ID', 'Sala', 'Usuario', 'Inicio', 'Fin', 'Proposito', 'Estado']]
    for r in reservas:
        data.append([
            r.id,
            r.sala.nombre,
            r.usuario.get_full_name() or r.usuario.username,
            timezone.localtime(r.fecha_inicio).strftime('%d/%m/%Y %H:%M'),
            timezone.localtime(r.fecha_fin).strftime('%d/%m/%Y %H:%M'),
            (r.proposito[:30] + '..') if len(r.proposito) > 30 else r.proposito,
            r.get_estado_display()
        ])

    # Tabla
    table = Table(data, colWidths=[30, 100, 120, 100, 100, 150, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#39A900")),  # Verde SENA
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


@login_required
@user_passes_test(es_admin)
def exportar_reservas_excel(request):
    """Genera un reporte Excel de todas las reservas."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Reservas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"

    # Encabezados
    headers = ['ID', 'Sala', 'Organizador', 'Propósito', 'Descripción', 'Asistentes', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Fecha Creación']
    ws.append(headers)

    # Estilo de encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="39A900", end_color="39A900", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Datos
    # Aplicar filtros si existen
    usuario_id = request.GET.get('usuario_id')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    reservas = Reserva.objects.all().select_related('sala', 'usuario')

    if usuario_id:
        reservas = reservas.filter(usuario_id=usuario_id)
    
    if fecha_inicio_str:
        f_inicio = parse_date(fecha_inicio_str)
        if f_inicio:
            dt_inicio = timezone.make_aware(datetime.combine(f_inicio, time.min))
            reservas = reservas.filter(fecha_inicio__gte=dt_inicio)
            
    if fecha_fin_str:
        f_fin = parse_date(fecha_fin_str)
        if f_fin:
            dt_fin = timezone.make_aware(datetime.combine(f_fin, time.max))
            reservas = reservas.filter(fecha_fin__lte=dt_fin)

    reservas = reservas.order_by('-fecha_inicio')

    for r in reservas:
        ws.append([
            r.id,
            r.sala.nombre,
            r.usuario.get_full_name() or r.usuario.username,
            r.proposito,
            r.descripcion or "",
            r.num_asistentes,
            timezone.localtime(r.fecha_inicio).strftime('%d/%m/%Y %H:%M'),
            timezone.localtime(r.fecha_fin).strftime('%d/%m/%Y %H:%M'),
            r.get_estado_display(),
            timezone.localtime(r.fecha_creacion).strftime('%d/%m/%Y %H:%M')
        ])

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response
